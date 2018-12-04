# utf-8
import json
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wiserd3.settings')
import django
django.setup()

from django.core.serializers import serialize
from openpyxl.utils import get_column_letter
from dataportal3.models import SpatialdataPostCodeS, SpatialdataPostCode
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.worksheet.read_only import ReadOnlyWorksheet
from dataportal3 import models as models
from django.db import connections
from djgeojson.serializers import Serializer as GeoJSONSerializer


class GEOIDtoGEOJSON:
    def __init__(self):
        pass

    def addGEOJSONfield(self, filename, sheet_name=None,
                        skip_rows=0, geo_id_header=None,
                        geo_id_column_number=0):

        print json.dumps(23.6754321)
        floatrepr = json.encoder.FLOAT_REPR
        print 'floatrepr', floatrepr

        # json.encoder.FLOAT_REPR = lambda o: format(o, '%.15g')
        print json.dumps(23.6754321)

        wb1 = load_workbook(filename)

        print wb1.get_sheet_names()

        if sheet_name:
            ws = wb1[sheet_name]
        else:
            ws = wb1[wb1.get_sheet_names()[0]]

        print type(ws)
        assert isinstance(ws, Worksheet)

        print 'This xls has {} rows'.format(ws.max_row)
        print 'This xls has {} columns'.format(ws.max_column)

        header_row = list(ws.iter_rows())[skip_rows]
        for cell in header_row:
            print cell.value

        new_column_letter = get_column_letter(ws.max_column + 1)
        new_column = ws[new_column_letter]
        geo_id_column_letter = get_column_letter(geo_id_column_number)

        # start_row = skip_rows + 1
        start_row = 8900
        end_row = 9506
        # end_row = 9300
        range_size = end_row - start_row

        new_column[skip_rows].value = 'GEOJSON'
        for row_number in range(start_row, end_row):

            geo_id_column_value = ws[geo_id_column_letter][row_number].value
            print '{}-{}-{}'.format(row_number, new_column_letter, geo_id_column_value)

            if geo_id_column_value:
                filter_var = 'label__istartswith'
                code = geo_id_column_value.replace(' ', '')

                print(code)
                postcode_subset = SpatialdataPostCode.objects.using('new').all().filter(**{filter_var: code})

                # TODO check if this needs better serialisation as below
                # s = serialize('geojson', postcode_subset, properties=('geom', 'label'))

                s = GeoJSONSerializer().serialize(
                    postcode_subset,
                    use_natural_keys=True,
                    with_modelname=False,
                    simplify=0.00005,
                    precision=5
                )
                print len(s), '\n'

                if len(s) > 32767:
                    s = GeoJSONSerializer().serialize(
                        postcode_subset,
                        use_natural_keys=True,
                        with_modelname=False,
                        simplify=0.00005,
                        precision=4
                    )
                    print len(s), ' After compression'

                    if len(s) > 32767:
                        print len(s), ' Still too big'
                        s = GeoJSONSerializer().serialize(
                            postcode_subset,
                            use_natural_keys=True,
                            with_modelname=False,
                            simplify=0.00005,
                            precision=3
                        )
                        if len(s) > 32767:
                            print len(s), ' Again still too big at 3sf'
                            s = None

                # new_column[row_number].value = '{}-{}'.format(row_number + 1, new_column_letter)
                new_column[row_number].value = s
                print '\n'

        wb1.save("test2.xlsx")


if __name__ == "__main__":

    filename = '/home/ianh/Downloads/UK-Mortgage-Lending2.xlsx'

    sr = GEOIDtoGEOJSON()
    sr.addGEOJSONfield(
        filename,
        sheet_name='UKF_time_series',
        skip_rows=1,
        geo_id_header='Sector',
        geo_id_column_number=1
    )
