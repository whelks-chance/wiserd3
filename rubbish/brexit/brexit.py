import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wiserd3.settings')
import django
django.setup()

from dataportal3 import models
import json
import requests
from openpyxl import Workbook


class BrexitRecording:
    def __init__(self):
        pass

    def setup(self):
        tmp_file = 'brexit_data.json'
        if os.path.exists(tmp_file):
            with open(tmp_file, 'r') as tmp_file_obj:
                res_obj = json.load(tmp_file_obj)
                print 'Loaded from file'
        else:
            url = 'https://public.broker.elections.api.bbci.co.uk/components/batch?components%5Beu_ref_banner_data_en%5D%5Bcomponent%5D=eu_ref_banner_data_en&components%5Beu_ref_ticker_en%5D%5Bcomponent%5D=eu_ref_ticker_en&components%5Bpolling_policy%5D%5Bcomponent%5D=polling_policy&components%5Beu_ref_nation_results_data_en%5D%5Bcomponent%5D=eu_ref_nation_results_data_en&components%5Beu_ref_latest_local_results_en%5D%5Bcomponent%5D=eu_ref_latest_local_results_en&components%5Beu_ref_map_data%5D%5Bcomponent%5D=eu_ref_map_data'

            res = requests.get(url)
            res_text = res.text
            print 'Loaded from URL'
            res_obj = json.loads(res_text)

            with open(tmp_file, 'a') as tmp_file_obj:
                tmp_file_obj.write(json.dumps(res_obj, indent=4))

        key_arr = [0, 2, 3, 5]

        # for c in res_obj['components']:
        for c in key_arr:
            print c

            try:
                c_body = json.loads(res_obj['components'][c]['body'])

                print c_body.keys()
            except:
                pass

        c_body = json.loads(res_obj['components'][5]['body'])

        filename = 'saved.xlsx'
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'region'
        ws.cell(row=1, column=2).value = 'leave_total'
        ws.cell(row=1, column=3).value = 'leave_percent'
        ws.cell(row=1, column=4).value = 'remain_total'
        ws.cell(row=1, column=5).value = 'remain_percent'
        ws.cell(row=1, column=6).value = 'outcome'
        ws.cell(row=1, column=7).value = 'outcome_int'
        ws.cell(row=1, column=8).value = 'turnout_percent'
        ws.cell(row=1, column=9).value = 'name'


        done_list = []
        itr_offset = 3
        for itr, key in enumerate(c_body.keys()):
            print '\n'
            # print c_body[key]
            itr_offset += 1

            ws.cell(row=itr_offset, column=1).value = key
            ws.cell(row=itr_offset, column=2).value = c_body[key]['leave']['total']
            ws.cell(row=itr_offset, column=3).value = c_body[key]['leave']['percent']
            ws.cell(row=itr_offset, column=4).value = c_body[key]['remain']['total']
            ws.cell(row=itr_offset, column=5).value = c_body[key]['remain']['percent']

            outcome = c_body[key]['outcome']
            outcome_int = None
            if outcome == 'L':
                outcome_int = 1
                outcome = 'Leave'
            if outcome == 'R':
                outcome_int = 2
                outcome = 'Remain'

            ws.cell(row=itr_offset, column=6).value = outcome
            ws.cell(row=itr_offset, column=7).value = outcome_int
            ws.cell(row=itr_offset, column=8).value = c_body[key]['turnout']['percent']
            ws.cell(row=itr_offset, column=9).value = c_body[key]['name']

            done_list.append(key)

            geom_obj = None
            try:
                # if key[:1] == 'W':
                geom_model = models.SpatialdataUKUA.objects.get(ctyua15cd=key)
                geom_obj = geom_model.geom
            except Exception as e1:
                try:
                    geom_model = models.SpatialdataUKLAD.objects.get(lad13cd=key)
                    geom_obj = geom_model.geom

                except Exception as e2:
                    try:
                        geom_model = models.SpatialdataUKLADSCO.objects.get(lad13cd=key)
                        geom_obj = geom_model.geom

                    except Exception as e3:
                        try:
                            geom_model = models.SpatialdataUA.objects.get(label=c_body[key]['name'])
                            geom_obj = geom_model.geom
                        except Exception as e4:
                            try:
                                geom_model = models.SpatialdataParlConstNI.objects.get(pc_id=key)
                                geom_obj = geom_model.geom

                            except Exception as e5:
                                print 'Exception', type(e5), e5, key

            br, created = models.Brexit.objects.get_or_create(region=key)
            # br.region = key
            br.leave_total = c_body[key]['leave']['total']
            br.leave_percent = c_body[key]['leave']['percent']
            br.remain_total = c_body[key]['remain']['total']
            br.remain_percent = c_body[key]['remain']['percent']
            br.outcome = outcome
            br.outcome_int = outcome_int
            br.turnout_percent = c_body[key]['turnout']['percent']
            br.name = c_body[key]['name']
            br.geom = geom_obj
            br.save()

        # print 'done_list', done_list
        return wb.save(filename)

if __name__ == "__main__":
    br = BrexitRecording()
    br.setup()