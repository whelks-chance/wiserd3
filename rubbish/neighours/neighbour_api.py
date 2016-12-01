# coding=utf-8
import os
from copy import deepcopy

from openpyxl.reader.excel import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wiserd3.settings')
import django
django.setup()
from django.db import connections
from dataportal3 import models

import json
import requests
from django.contrib.gis.geos import GEOSGeometry
from openpyxl import Workbook


class Neighbours():
    def __init__(self):
        self.area_ids = set()
        self.table_as_arrays = []

        self.default_geojson = { "type": "FeatureCollection",
                                 "features": [
                                 ]
                                 }
        self.default_feature = { "type": "Feature",
                                 "geometry": {
                                     "type": "Polygon",
                                     "coordinates": [
                                         [ [100.0, 0.0], [101.0, 0.0], [101.0, 1.0],
                                           [100.0, 1.0], [100.0, 0.0] ]
                                     ]
                                 },
                                 "properties": {
                                     "prop0": "value0",
                                     "prop1": {"this": "that"}
                                 }
                                 }

    def do_thing(self, lat='51.4791121', lng='-3.1779989'):

        api_url = "https://api.neighbourhoodalert.co.uk/api/NHWNScheme/getLocalSchemes?lat={}&lng={}&numberPerPage=100&page=1".format(lat, lng)

        page_url = "http://www.ourwatch.org.uk/postcode-search/?postcode=cf101aa"

        res = requests.get(
            api_url,
            headers={
                'Authorization': 'Bearer AP__ApOeCABbigwm2NsBRQJlJr9dKBzbwTQch3OtrQGPOiL97ccoeZYir07kG9IgTtFnZNFNOHAuuXiOzbhKPC1YagS1EOL5eDy9JkyTZd_0lOlDCgzZwmhkzm9xzFeoLkMKHzncF3SLC4DLHAvH5JxUcWTawHWE1mHuHRyeBx-sfalfWMlrMPxJMiplao5vAxUnviBucnqmsSMC-pgylvqs1c4jNNisGzjF4GH3mfMPf1NOXwY6dDExr_xVNREDxnlVq5fpgII4n6LreZ5Qf4WMGeq6VGerwZElDBuZZ4xO0bjVns0nPd9m0MJr3jTqcnxLPbtyZLgm0XRzh45JxliEsmvofKIv1uJvYLcMjj6GeQ-q6EP4nj-gOqJ2rdYCTKuV3fJ0FRW2XIkpfwWXgPjxceWRIj-VSlsilLzrXJg'
            }
        )
        # print res.text

        json_res = json.loads(res.text)

        for a in json_res:
            print a['scheme_name']
            nhw_scheme_id = a['nhw_scheme_id']
            num_members = a['scheme_number_of_members']
            print nhw_scheme_id

            if nhw_scheme_id not in self.area_ids:
                self.area_ids.add(nhw_scheme_id)

                wkt = a['area']
                if 'MULTI' not in wkt:
                    if 'LINESTRING' in wkt:
                        pass
                    else:
                        wkt = 'LINESTRING (' + wkt + ')'
                else:
                    wkt = wkt.replace('(', '')
                    wkt = wkt.replace(')', '')
                    wkt = wkt.replace('MULTI', 'LINESTRING(')
                    wkt += ')'

                try:
                    geom = GEOSGeometry(wkt, srid=4326)

                    # print geom.json
                    convex_hull = geom.convex_hull
                    # print convex_hull
                    print convex_hull.area
                    # print convex_hull.centroid
                    print convex_hull.json

                    self.table_as_arrays.append([
                        nhw_scheme_id,
                        a['scheme_name'],
                        a['postcode'],
                        num_members,
                        a['number_of_households'],
                        convex_hull.centroid.json,
                        wkt,
                        convex_hull.json
                    ])

                    print '\n\n'
                except Exception as err:
                    print err, nhw_scheme_id

            else:
                print 'Repeated', nhw_scheme_id

        print len(json_res)
        print self.area_ids

    def save(self, filename='neighbourhood_lists.xls', skip=[]):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'id'
        ws.cell(row=1, column=2).value = 'name'
        ws.cell(row=1, column=3).value = 'postcode'

        ws.cell(row=1, column=4).value = 'scheme_number_of_members'
        ws.cell(row=1, column=5).value = 'number_of_households'
        ws.cell(row=1, column=6).value = 'centroid'
        ws.cell(row=1, column=7).value = 'WKT'
        ws.cell(row=1, column=8).value = 'JSON'

        done_list = []
        for itr, school in enumerate(self.table_as_arrays):

            if school not in skip:
                itr_offset = itr + 3

                # print type(table_as_arrays[itr][0][0])

                ws.cell(row=itr_offset, column=1).value = unicode(self.table_as_arrays[itr][0])
                ws.cell(row=itr_offset, column=2).value = unicode(self.table_as_arrays[itr][1])
                ws.cell(row=itr_offset, column=3).value = unicode(self.table_as_arrays[itr][2])
                ws.cell(row=itr_offset, column=4).value = unicode(self.table_as_arrays[itr][3])
                ws.cell(row=itr_offset, column=5).value = unicode(self.table_as_arrays[itr][4])
                ws.cell(row=itr_offset, column=6).value = unicode(self.table_as_arrays[itr][5])
                ws.cell(row=itr_offset, column=7).value = unicode(self.table_as_arrays[itr][6])
                ws.cell(row=itr_offset, column=8).value = unicode(self.table_as_arrays[itr][7])

                done_list.append(school)
            else:
                print 'Skipped {}'.format(school)

        # print 'done_list', done_list
        return wb.save(filename)

    def build_grid(self):

        # Get all the postcodes

        districts = ['LL58', 'LL59', 'LL60', 'LL61', 'LL62', 'LL64', 'LL65', 'LL66', 'LL67', 'LL68', 'LL69', 'LL70', 'LL71', 'LL72', 'LL73', 'LL74', 'LL75', 'LL76', 'LL77', 'LL78',
                     'NP1', 'NP2', 'NP23', 'NP3',
                     'CF31', 'CF32', 'CF33', 'CF34', 'CF35', 'CF36',
                     'CF3', 'CF46', 'CF81', 'CF82', 'CF83', 'NP1', 'NP2',
                     'CF1', 'CF2', 'CF3', 'CF4', 'CF5', 'CF83',
                     'SA14', 'SA15', 'SA16', 'SA17', 'SA18', 'SA19', 'SA20', 'SA31', 'SA32', 'SA33', 'SA34', 'SA35', 'SA38', 'SA39', 'SA4', 'SA40', 'SA44', 'SA48', 'SA66',
                     'SA38', 'SA40', 'SA43', 'SA44', 'SA45', 'SA46', 'SA47', 'SA48', 'SY20', 'SY23', 'SY24', 'SY25',
                     'LL16', 'LL18', 'LL21', 'LL22', 'LL24', 'LL25', 'LL26', 'LL27', 'LL28', 'LL29', 'LL30', 'LL31', 'LL32', 'LL33', 'LL34', 'LL57',
                     'CH7', 'LL11', 'LL15', 'LL16', 'LL17', 'LL18', 'LL19', 'LL20', 'LL21', 'LL22',
                     'CH1', 'CH4', 'CH5', 'CH6', 'CH7', 'CH8', 'LL11', 'LL12', 'LL18', 'LL19',
                     'CF1', 'CF32', 'CF35', 'CF5', 'CF61', 'CF62', 'CF63', 'CF64', 'CF71',
                     'LL21', 'LL23', 'LL33', 'LL35', 'LL36', 'LL37', 'LL38', 'LL39', 'LL40', 'LL41', 'LL42', 'LL43', 'LL44', 'LL45', 'LL46', 'LL47', 'LL48', 'LL49', 'LL51', 'LL52', 'LL53', 'LL54', 'LL55', 'LL56', 'LL57', 'SY20',
                     'CF46', 'CF47', 'CF48',
                     'NP4', 'NP5', 'NP6', 'NP7',
                     'SA10', 'SA11', 'SA12', 'SA13', 'SA18', 'SA8', 'SA9',
                     'CF3', 'NP1', 'NP10', 'NP19', 'NP20', 'NP6', 'NP9',
                     'SA34', 'SA35', 'SA36', 'SA37', 'SA41', 'SA42', 'SA43', 'SA61', 'SA62', 'SA63', 'SA64', 'SA65', 'SA66', 'SA67', 'SA68', 'SA69', 'SA70', 'SA71', 'SA72', 'SA73',
                     'CF44', 'CF48', 'HR3', 'HR5', 'LD1', 'LD2', 'LD3', 'LD4', 'LD5', 'LD6', 'LD7', 'LD8', 'NP7', 'NP8', 'SA10', 'SA11', 'SA9', 'SY10', 'SY15', 'SY16', 'SY17', 'SY18', 'SY19', 'SY20', 'SY21', 'SY22', 'SY5',
                     'CF35', 'CF37', 'CF38', 'CF39', 'CF4', 'CF40', 'CF41', 'CF42', 'CF43', 'CF44', 'CF45', 'CF72',
                     'SA1', 'SA18', 'SA2', 'SA3', 'SA4', 'SA5', 'SA6', 'SA7',
                     'NP4', 'NP44', 'NP6',
                     'LL11', 'LL12', 'LL13', 'LL14', 'LL20', 'SY13', 'SY14']

        print len(districts)

        for pcode in districts:
            print pcode
            first_found = models.SpatialdataPostCode.objects.filter(label__startswith=pcode)[:1]
            if len(first_found):
                print first_found[0].label
                try:
                    geom = GEOSGeometry(first_found[0].geom, srid=27700)
                    coords = json.loads(geom.centroid.json)['coordinates']
                    lat = coords[1]
                    lng = coords[0]
                    print lat, lng
                    print geom.area
                    print '\n'
                    self.do_thing(lat, lng)
                    self.save(filename='neighbourhood_lists3_{}.xls'.format(pcode))
                except Exception as e2:
                    print e2, pcode, 'DB Error'

        nw_x = -5.60852
        nw_y = 53.48805
        se_x = -2.00000
        se_y = 51.13800

    def read_xls_build_geojson(self):
        wb2 = load_workbook('neighbourhood_lists3.xlsx')

        print wb2.get_sheet_names()

        ws = wb2[wb2.get_sheet_names()[0]]
        colJSON = ws['H']
        colA = ws['A']
        colB = ws['B']
        colC = ws['C']
        colD = ws['D']
        colE = ws['E']

        print len(colJSON)

        all_features = []
        for itr in range(2, len(colJSON)):
            print colJSON[itr].value

            new_feature = deepcopy(self.default_feature)

            new_feature['geometry'] = json.loads(colJSON[itr].value)
            geom = GEOSGeometry(colJSON[itr].value)

            new_feature["properties"] = {
                "id": colA[itr].value,
                "name": colB[itr].value,
                "postcode": colC[itr].value,
                "scheme_number_of_members": colD[itr].value,
                "number_of_households": colE[itr].value,
                "area": geom.area

            }
            new_feature["crs"] = {
                'type': 'name',
                'properties': {
                    'name': 'WGS84'
                }
            }

            all_features.append(new_feature)

        self.default_geojson['features'] = all_features

        print self.default_geojson

        with open('big.geojson', 'wr') as all_features_file:
            all_features_file.write(json.dumps(self.default_geojson, indent=4))


if __name__ == "__main__":
    nwapi = Neighbours()
    # nwapi.build_grid()
    # nwapi.save(filename='neighbourhood_lists3.xls')
    # nwapi.do_thing()
    nwapi.read_xls_build_geojson()
