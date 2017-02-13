import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "wiserd3.settings")
import django
django.setup()

from dataportal3 import models
import json
import pprint
import xml.etree.ElementTree
import requests
from openpyxl import Workbook


class FreeSchoolMeals():
    def __init__(self):
        self.school_data = None

    def prepare(self):

        # http://mylocalschool.wales.gov.uk/Data/schools.json?111111
        # http://mylocalschool.wales.gov.uk/schools.json?{}
        res = requests.get('http://mylocalschool.wales.gov.uk/Data/schools.json?111111')
        print 'Retrieved school data'
        self.school_data = json.loads(res.text)
        print 'Found {} schools ({} reported)'.format(len(self.school_data['schools']), len(self.school_data['schools']))

    def get_school_data(self, limit=2):
        data = {}
        with open('temp.txt', 'a') as tmp_file:

            save_count = 0
            for school in self.school_data['schools'][:limit]:
                try:
                    school_code = school['schoolCode']
                    res = requests.get(
                        'http://mylocalschool.wales.gov.uk/schools/{}.json?807022358'.format(school_code)
                    )
                    print 'Retrieved school data for {}.'.format(school_code)
                    school_detail = json.loads(res.text)

                    with open('school_{}_data.txt'.format(school_code), 'a') as school_tmp_file:
                        school_tmp_file.write(json.dumps(school_detail, indent=4))

                    school_detail_summary = school_detail['lstSummary']

                    school_detail_summary_dict = {}
                    for idx, item in enumerate(school_detail_summary):
                        # print idx
                        # print item
                        school_detail_summary_dict[str(idx)] = item

                    school_model = models.SchoolData()

                    school_model.lat = school['lat']
                    school_model.lng = school['lng']

                    school_model.lea = school['lea']
                    school_model.name = school['name']
                    school_model.schoolCode = school['schoolCode']
                    school_model.schoolType = school['schoolType']
                    school_model.statusNameEng = school['statusNameEng']
                    school_model.statusNameWelsh = school['statusNameWelsh']

                    school_model.LEANameEnglish = school_detail['LEANameEnglish']
                    school_model.LEANameWelsh = school_detail['LEANameWelsh']
                    school_model.schTypeEnglish = school_detail['schTypeEnglish']
                    school_model.schTypeWelsh = school_detail['schTypeWelsh']
                    school_model.schLanguageEnglish = school_detail['schLanguageEnglish']
                    school_model.schLanguageWelsh = school_detail['schLanguageWelsh']
                    school_model.genderMix = school_detail['genderMix']
                    school_model.commentsEng = school_detail['commentsEng']
                    school_model.commentsCym = school_detail['commentsCym']
                    school_model.schoolGroupingCodePrim = school_detail['schoolGroupingCodePrim']
                    school_model.addressLine1 = school_detail['addressLine1']
                    school_model.addressLine2 = school_detail['addressLine2']
                    school_model.addressLine3 = school_detail['addressLine3']
                    school_model.addressLine4 = school_detail['addressLine4']
                    school_model.postcode = school_detail['postcode']
                    school_model.telephoneNo = school_detail['telephoneNo']

                    found_postcodes_trimmed = models.SpatialdataPostCodePoint.objects.filter(postcode=school_detail['postcode'].replace(' ', ''))
                    if found_postcodes_trimmed.count() > 0:
                        found_postcode_geom = found_postcodes_trimmed[0].geom
                        school_model.geom = found_postcode_geom
                    else:
                        found_postcodes_orig = models.SpatialdataPostCodePoint.objects.filter(
                            postcode=school_detail['postcode'])
                        if found_postcodes_orig.count() > 0:
                            found_postcode_geom = found_postcodes_orig[0].geom
                            school_model.geom = found_postcode_geom
                        else:
                            print '\n', 'err ***', school_code, '\n'
                            print 'No postcode for ', school, school_detail['postcode']

                    school_model.school_dict = self.flatten_dict(school_detail_summary_dict)
                    school_model.school_json = school_detail_summary

                    school_model.save()

                    save_count += 1
                    print 'Saved : ', save_count, school_detail['postcode']

                    # address_line = school_detail['school']['basicDetails']['address']
                    # address_line = '<a>' + address_line + '</a>'
                    # e = xml.etree.ElementTree.fromstring(address_line)
                    # postcode = e.findall('li')[-1].text
                    #
                    # data[school_code] = {
                    #     'schoolCode': school['schoolCode'],
                    #     'name': school['name'],
                    #     'fsm': school_detail['school']['pupilPopulation']['fsm']['data'],
                    #     'postcode': postcode
                    # }
                    #
                    # tmp_file.write(
                    #     pprint.pformat(data[school_code]) + ',\n'
                    # )

                except Exception as e:
                    print '\n', 'err ***', '\n'
                    print e, type(e), school
                    print '\n', 'err ***', '\n'

        return data

    def flatten_dict(self, d):

        def expand(key, value):
            if isinstance(value, dict):
                return [ (key + '.' + k, v) for k, v in self.flatten_dict(value).items()]
            else:
                return [ (key, value) ]

        items = [ item for k, v in d.items() for item in expand(k, v) ]

        return dict(items)

    def get_school_data_xls(self, filename='default.xls', limit=2, skip=None):
        if not skip:
            skip = []

        all_school_data = self.get_school_data(limit=limit)

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'name'
        ws.cell(row=1, column=2).value = 'postcode'
        ws.cell(row=1, column=3).value = 'fsm'

        done_list = []
        for itr, school in enumerate(all_school_data):

            if school not in skip:
                itr_offset = itr + 2
                # print all_school_data[school]

                ws.cell(row=itr_offset, column=1).value = all_school_data[school]['name']
                ws.cell(row=itr_offset, column=2).value = all_school_data[school]['postcode']
                ws.cell(row=itr_offset, column=3).value = all_school_data[school]['fsm'][-1]

                done_list.append(school)
            else:
                print 'Skipped {}'.format(school)

        print 'done_list', done_list
        return wb.save(filename)

    def find_school(self):
        founds = models.SchoolData.objects.filter(postcode='LD3 9SR')
        for fi in founds:
            print fi.__dict__, '\n\n**'

        found_postcodes1 = models.SpatialdataPostCodePoint.objects.filter(postcode='LD3 9SR'.replace(' ', ''))
        if found_postcodes1.count() == 0:
            found_postcodes2 = models.SpatialdataPostCodePoint.objects.filter(postcode='LD3 9SR')
            if found_postcodes2.count() == 0:
                print 'super fail'
            else:
                print found_postcodes2, 'LD3 9SR'
        else:
            print found_postcodes1, 'LD3 9SR'.replace(' ', '')

    def fix_null_geom(self):
        null_geoms = models.SchoolData.objects.filter(geom=None)
        print null_geoms.count()

        for ng in null_geoms:
            assert isinstance(ng, models.SchoolData)

            found_postcodes_trimmed = models.SpatialdataPostCodePoint.objects.filter(
                postcode=ng.postcode.replace(' ', ''))
            if found_postcodes_trimmed.count() > 0:
                found_postcode_geom = found_postcodes_trimmed[0].geom
                ng.geom = found_postcode_geom
            else:
                found_postcodes_orig = models.SpatialdataPostCodePoint.objects.filter(
                    postcode=ng.postcode)
                if found_postcodes_orig.count() > 0:
                    found_postcode_geom = found_postcodes_orig[0].geom
                    ng.geom = found_postcode_geom
                    print 'found postcode for ', ng.postcode
                else:
                    print '\n', 'err ***', ng.schoolCode, '\n'
                    print 'No postcode for ', ng.__dict__, ng.postcode

            ng.save()

    def find_fsm(self):
        schools = models.SchoolData.objects.all()

        count = 0
        count2 = 0
        for school in schools:
            print '\n'
            assert isinstance(school, models.SchoolData)
            school_dict = school.school_dict

            items = [(key, value) for key, value in school_dict.items() if value and 'fsm' in value.lower()]

            if(items):
                print items

                for item in items:
                    print item[0]

                    fsm_idx = item[0].split('.')[0]
                    print fsm_idx

                    if fsm_idx != '4':
                        print '********************'
                        print school.postcode

                    score_key = fsm_idx + '.score'

                    print school_dict[score_key]
                    school.fsm_percent = school_dict[score_key]
                    school.save()
                    count += 1

                count2 += len(items)

        print count2
        print count

if __name__ == "__main__":
    fsm = FreeSchoolMeals()

    # flat_D = fsm.flatten_dict({
    #     'a': {
    #         'b': [1,2,3,4, 'asdfg']
    #     },
    #     'c': 5673257632,
    #     'd': [1,2,3,4,5, {
    #         'ab': 'cd'
    #     }]
    # })

    # print flat_D

    # fsm.prepare()
    # fsm.get_school_data_xls(
    #     filename='school_data_all.xlsx',
    #     limit=2000,
    #     skip=[]
    # )

    # fsm.fix_null_geom()

    # fsm.find_fsm()

    # fsm.find_school()
