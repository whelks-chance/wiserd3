import json
import pprint

from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell


class RawQuestion:
    def __init__(self, variable_name):
        self.question_number = ''
        self.response_table = ''
        self.question_text = ''
        self.variable_name = variable_name

    def description(self):
        return {
            'variableid': self.variable_name,
            'question_text': self.question_text,
            'response_table': self.response_table,
            'question_number': self.question_number
        }


class Sample:
    def __init__(self):
        self.questions = []

    def samples_questions(self):
        return self.questions

    def description(self):
        descriptions = []
        for q in self.questions:
            descriptions.append(q.description())
        return {'questions': descriptions}

letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm',
           'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']


class QuestionResponseLinker:
    def __init__(self, qbank_file, sweep_cohort_responses_file, variable_name_filter_str):
        self.variable_name_filter_str = variable_name_filter_str
        self.sweep_cohort_responses_file = sweep_cohort_responses_file
        self.qbank_file = qbank_file

    def run(self):
        qbank_wb = load_workbook(self.qbank_file)
        print qbank_wb.get_sheet_names()
        ws1 = qbank_wb[qbank_wb.get_sheet_names()[0]]
        assert isinstance(ws1, Worksheet)

        scres_wb = load_workbook(self.sweep_cohort_responses_file)
        print scres_wb.get_sheet_names()
        ws2 = scres_wb[scres_wb.get_sheet_names()[0]]
        assert isinstance(ws2, Worksheet)

        sample = Sample()

        question_number = 1
        question_letter_itr = 0

        variable_name_col = ws1['E']
        response_value_map_col =ws1['D']
        question_text_col = ws1['C']

        # For each row in Question Bank, get the question identifier "variable name"
        for variable_name_cell in variable_name_col[1:]:
            assert isinstance(variable_name_cell, Cell)

            if not self.variable_name_filter_str in variable_name_cell.value == None:
                # if self.variable_name_filter_str in variable_name_cell.value is None:
                #     continue
                print('We are on row {}'.format(variable_name_cell.row))

                rq = RawQuestion(variable_name_cell.value)

                # codings for responses e.g. 1 'Yes' 2 'No'
                mapping_cell = response_value_map_col[variable_name_cell.row-1]
                mapping_dict = {}

                print ('Mapping cell value', mapping_cell.value)
                if mapping_cell.value == '0' or mapping_cell.value == 0:
                    print(type(mapping_cell.value))
                    print('Zero mapping')
                else:
                    if not mapping_cell.value == '0':
                        # Only populate this if the questionbank file says how

                        # Turn 1 'Yes' 2 'No' into
                        # 1 'Yes'
                        # 2 'No'
                        print('MAPPING CELL VALUE', mapping_cell.value)
                        mapping_list = str(mapping_cell.value).encode('utf-8').split("' ")
                        for map in mapping_list:
                            map_text_value = map.split(' ')
                            # 1 'Yes' into
                            # 1
                            # 'Yes'
                            print(map_text_value)
                            # produces Yes without speech marks or spaces as the key, and the number as the value.
                            # {
                            #     'Yes': 1
                            # }
                            mapping_dict[
                                    " ".join(map_text_value[1:]).lstrip("'").rstrip("'")
                                ] = map_text_value[0].strip()

                    print(mapping_dict)

                question_text = question_text_col[variable_name_cell.row-1]
                print(question_text.value)

                # Add question text to RawQuestion object defined above
                rq.question_text = question_text.value

                # question_number_str = str(question_number)
                # if ':' in rq.question_text:
                #     if question_letter_itr == 0:
                #         question_number += 1
                #     question_number_str += letters[question_letter_itr]
                #     question_letter_itr += 1
                # else:
                #     question_number += 1
                #     question_number_str = str(question_number)
                #
                #     question_letter_itr = 0

                # print('*********************************\n\n')

                for row in ws2.iter_rows(min_row=1, max_row=1):
                    for idx, cell in enumerate(row):

                        # For each column header in the first row of results xlsx, look for the variable_name above
                        # Question answers in the redacted responses file are by column
                        if cell.value == variable_name_cell.value:

                            print('The response xlsx column number {} labelled {} is for question {}'.format(
                                idx, cell.column, variable_name_cell.value))

                            print(idx, cell.value)
                            assert isinstance(cell, Cell)

                            rq.question_number = idx

                            # print(cell.col_idx)
                            # print(cell.column)

                            results_and_freqs = {}

                            # Once we have the right cell in the column headers
                            # Get the column for that header, move down the column
                            for cell in ws2[cell.column][1:]:
                                # print(cell.value)

                                # If we've seen this response before, +1 to the counter
                                if cell.value in results_and_freqs:
                                    results_and_freqs[cell.value] += 1
                                else:
                                    # New response, add it to the dict, set count to 1
                                    results_and_freqs[cell.value] = 1

                            print('Question {} has result freq :'.format(variable_name_cell.value))
                            print pprint.pformat(results_and_freqs)

                            response_table = []
                            # we didn't populate the mapping dict, as it was 0 in the question bank file
                            if len(mapping_dict):
                                # Build the response table
                                for response in results_and_freqs:
                                    response_detail = {
                                        'response': response,
                                        'response (n)': results_and_freqs[response]
                                    }
                                    if response == 'Missing':
                                        response_detail['value'] = '99'

                                    if response in mapping_dict:
                                        response_detail['value'] = mapping_dict[response]

                                    response_table.append(response_detail)
                            print('\n')
                            print pprint.pformat(response_table)

                            # Add response table to RawQuestion object
                            rq.response_table = response_table

                            print(rq.__dict__)

                            # Add question to the sample bank and move on
                            sample.questions.append(rq)

                            print('****\n\n')

        return sample



                        # found = True
                # if not found:
                #     print('Didnt find {}'.format(variable_name_cell.value))


if __name__ == '__main__':
    qbank_file = '/home/kdickson/QuestionBank_Education.xlsx'
    sweep_cohort_responses_file = '/home/kdickson/Surveys/Sweep8_Redacted.xlsx'

    qrl = QuestionResponseLinker(qbank_file, sweep_cohort_responses_file, '8')

    sample = qrl.run()
    with open('sample_8.json', 'w') as outputfile:
        outputfile.write(json.dumps(sample.description(), indent=4))
