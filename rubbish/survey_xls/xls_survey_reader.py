# utf-8
import json
import os

import datetime

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wiserd3.settings')
import django
django.setup()

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.worksheet.read_only import ReadOnlyWorksheet
from dataportal3 import models as models
from django.db import connections


class SurveyReader:
    def __init__(self, filename):
        self.filename = filename
        self.surveys = {}
        self.dcs = {}

    def get_dc_and_survey(self, sweep, cohort):

        title = 'WISERD Education {} {}'.format(sweep, cohort)

        dc_identifier = 'dc_{}'.format(title.replace(' ', '_').lower())
        dc_title = title
        survey_title = title

        new_dc, created = models.DcInfo.objects.get_or_create(identifier=dc_identifier)
        if created:
            new_dc.title = dc_title
            new_dc.type = models.DublincoreType.objects.get(dc_type_title='Text')
            new_dc.format = models.DublincoreFormat.objects.get(dc_format_title='text')
            new_dc.save()

            self.dcs[(sweep, cohort)] = new_dc

            print new_dc

        new_survey, created = models.Survey.objects.get_or_create(surveyid=new_dc.identifier)

        if created:
            new_survey.identifier = new_survey.surveyid
            new_survey.dublin_core = new_dc
            new_survey.survey_title = survey_title
            new_survey.save()

            self.surveys[(sweep, cohort)] = new_survey

            print new_survey

        return new_dc, new_survey

    def read_and_write(self):
        filename1 = self.filename
        wb1 = load_workbook(filename1)

        print wb1.get_sheet_names()

        ws = wb1[wb1.get_sheet_names()[0]]

        print type(ws)

        assert isinstance(ws, Worksheet)

        colA = ws['A']
        colB = ws['B']
        colC = ws['C']
        colD = ws['D']
        colQ = ws['Q']
        colX = ws['X']

        num_questions = len(ws['A']) - 1

        print 'This xls has {} questions'.format(num_questions)

        print colA[0].value, colB[0].value, colC[0].value

        for i in range(1, num_questions):
            q_sweep = colA[i].value
            q_cohort = colB[i].value
            q_name = colD[i].value
            q_label = colQ[i].value
            q_response_table = colX[i].value

            if q_label and len(q_label.strip()):

                new_dc, new_survey = self.get_dc_and_survey(q_sweep, q_cohort)

                fullname = '{}_{}_{}'.format(q_sweep, q_cohort, q_name)
                print 'Question {} has question text {}'.format(fullname.encode('utf-8'), q_label.encode('utf-8'))

                qid = '{}_{}'.format(new_survey.surveyid, fullname)
                new_q, created = models.Question.objects.get_or_create(qid=qid, survey=new_survey)
                # new_q.survey = new_survey
                new_q.questionnumber = q_name
                new_q.literal_question_text = q_label
                new_q.variableid = q_name

                new_response = models.Response()
                new_response.responseid = 'res_{}'.format(new_q.qid)
                # new_response.responsetext
                new_response.save()

                new_q.response = new_response
                new_q.save()

                if q_response_table and len(q_response_table.strip()):

                    new_response.response_type = models.ResponseType.objects.get(response_name='Closed-ended CATEGORICAL')
                    new_response.save()

                    res_table_arr = []
                    print q_response_table

                    q_res_split = q_response_table.split('\' ')

                    for a in q_res_split:
                        # print a

                        idx = a.find('\'')
                        a_key = a[:idx]
                        a_cat = a[idx:]
                        # print a_key, ":", a_cat

                        option = {
                            'category': a_cat.rstrip('\'').lstrip('\'').encode('utf-8'),
                            'value': a_key.encode('utf-8')
                        }
                        res_table_arr.append(option)

                        print '\n'

                    print res_table_arr
                    new_response_table, created = models.ResponseTable.objects.get_or_create(response=new_response)
                    # new_response_table.response = new_response
                    new_response_table.feature_attributes = res_table_arr
                    new_response_table.save()
                else:
                    new_response.response_type = models.ResponseType.objects.get(response_name='Open-ended QUALITATIVE')
                    new_response.save()

                print 'Saved question {}'.format(i)

        # print '', connections['new'].queries
        print self.surveys
        print self.dcs

    def build_from_description_file(self, survey_description_file, sweep, cohort):
        dc, survey = self.get_dc_and_survey(sweep, cohort)

        with open(survey_description_file, 'r') as f1:
            json_blob = json.load(f1)

            for q in json_blob['questions']:
                print(q)
                self.create_question(survey, q)

    def create_question(self, survey, raw_question_data):

        fullname = raw_question_data['variableid'] + str(raw_question_data['question_number'])

        qid = '{}_{}'.format(survey.surveyid, fullname)
        new_q, created = models.Question.objects.get_or_create(qid=qid, survey=survey)
        # new_q.survey = new_survey
        new_q.questionnumber = raw_question_data['question_number']
        new_q.literal_question_text = raw_question_data['question_text']
        new_q.variableid = raw_question_data['variableid']
        new_response = models.Response()
        new_response.responseid = 'res_{}'.format(new_q.qid)
        # new_response.responsetext
        new_response.save()

        new_q.response = new_response
        new_q.save()

        q_response_table = raw_question_data['response_table']
        if q_response_table is not None:

            if len(q_response_table):
                new_response.response_type = models.ResponseType.objects.get(response_name='Closed-ended CATEGORICAL')
            else:
                new_response.response_type = models.ResponseType.objects.get(response_name='Open-ended QUALITATIVE')
            new_response.save()

            new_response_table, created = models.ResponseTable.objects.get_or_create(response=new_response)
            # new_response_table.response = new_response
            new_response_table.feature_attributes = q_response_table
            new_response_table.save()
        else:
            print('No response table found')
            new_response.response_type = models.ResponseType.objects.get(response_name='Open-ended QUALITATIVE')
            new_response.save()

        print 'Saved question {}'.format(new_q)


if __name__ == "__main__":

    # /home/ianh/Downloads/educationdata.xlsx
# /home/ianh/Downloads/sample_4C.json

    # filename = '/home/kdickson/text_survey.xlsx'
    filename = '/home/kdickson/educationdata.xlsx'
    # filename = '/home/kdickson/educationdata_small.xlsx'

    sr = SurveyReader(filename)
    # sr.read_and_write()
    sr.build_from_description_file(
        '/home/kdickson/Downloads/sample_1C.json',
        '1',
        'C'

    )
