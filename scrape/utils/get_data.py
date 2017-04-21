# coding=utf-8
import json
import os

from django.contrib.gis.geos import Point

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wiserd3.settings')
import django
django.setup()

import requests
# from bs4 import BeautifulSoup
from BeautifulSoup import BeautifulSoup
from openpyxl import Workbook
from scrape import models
from dataportal3 import models as dataportal_models



class TaxService:
    def __init__(self):
        self.base_url = 'https://www.tax.service.gov.uk/view-my-valuation/list-valuations-advanced?primaryCriteria=BA&number=&street=&town=&postCode=&specialCategoryCode=&baRef=&descriptionCode=&from=&to={}{}#search-results'
        self.billing_authority_code = '6815'
        self.billing_authority_name = 'Cardiff'
        # self.billing_auth = '&billingAuthority={}'.format(self.billing_authority_code)
        self.start_page = '&startPage=1'

    def get_table_data(self, page_soup):
        table_as_arrays = []
        try:
            table = page_soup.find(name='table', attrs={'id': 'search-results-table'})
            table_body = table.find(name='tbody')
            table_rows = table_body.findAll(name='tr')

            for table_row in table_rows:
                row_array = []
                table_cells = table_row.findAll('td')
                for cell in table_cells:

                    if hasattr(cell.span, 'a') and cell.span.a:
                        print (cell.span.a.contents)
                        row_array.append(cell.span.a.contents)
                    else:
                        print (cell.span.contents)
                        row_array.append(cell.span.contents)
                table_as_arrays.append(row_array)
                print ('****\n\n')

            print (len(table_rows))
        except:
            print('Error reading page *****')
        return table_as_arrays

    def do_thing(self):
        table_as_arrays = []
        url = self.base_url.format(
            '&billingAuthority={}'.format(self.billing_authority_code),
            self.start_page)
        response = requests.get(url)

        soup = BeautifulSoup(response.text, 'html.parser')
        table_as_arrays.extend(self.get_table_data(soup))

        with open('tmp_{}_{}_{}_properties.json'.format(
                self.billing_authority_name,
                self.billing_authority_code,
                '1'
        ), 'a') as tmp:
            tmp.write(json.dumps(table_as_arrays, indent=4))

        # result-count

        search_results = soup.find(name='span', attrs={'id': 'result-count'})
        # print(search_results.__dict__)

        print('***')
        records = search_results.contents[0].split('of')[1].replace(',', '').strip()
        print(records)
        # print(int(records))
        # print(int(records) / 25)

        pages_total = int(int(records) / 25) + 1
        print(pages_total)

        # import sys
        # sys.exit()

        # search_results = soup.find(name='section', attrs={'id': 'search-results'})
        # footer = search_results.find(name='footer')
        # ul = footer.find(name='ul')
        # lis = ul.findAll(name='li')

        # for li in lis[:-1]:
        #     if hasattr(li, 'a') and li.a:
        #         697
        for page_counter in range(2, pages_total):
            print(self.billing_authority_name, page_counter, ' / ', pages_total)

            # print('Doing page : ', li.a.contents[0].strip())
            # print('contents', li.a.contents)
            # print(li.a.attrs)
            # print (li.a.attrs[0][1])

            # page_itr = li.a.contents[0].strip()

            url = self.base_url.format(
                '&billingAuthority={}'.format(self.billing_authority_code),
                '&startPage=' + str(page_counter)
            )
            response = requests.get(url)

            soup = BeautifulSoup(response.text)
            page_table = self.get_table_data(soup)

            table_as_arrays.extend(page_table)

            print ('\n')

            directory = '{}_{}'.format(self.billing_authority_name, self.billing_authority_code)
            if not os.path.exists(directory):
                os.makedirs(directory)

            tmpfile = 'tmp_{}_{}_{}_properties.json'.format(
                    self.billing_authority_name,
                    self.billing_authority_code,
                    str(page_counter)
            )

            tmpfile_in_dir = os.path.join(directory, tmpfile)

            with open(tmpfile_in_dir, 'a') as tmp:
                tmp.write(json.dumps(page_table, indent=4))

        self.save(
            table_as_arrays,
            filename='tax_lists__{}_{}_billing_authority.xls'.format(
                self.billing_authority_name,
                self.billing_authority_code
            )
        )

    def save(self, table_as_arrays, filename='tax_lists1.xls', skip=[]):
        wb = Workbook()
        ws = wb.active

        # Address
        # Description
        # Total area(m² / unit)
        # Current rateable value

        ws.cell(row=1, column=1).value = 'Address'
        ws.cell(row=1, column=2).value = 'Description'
        ws.cell(row=1, column=3).value = 'Total area(m² / unit)'
        ws.cell(row=1, column=4).value = 'Price per(m² / unit)'
        ws.cell(row=1, column=5).value = 'Current rateable value'

        done_list = []
        for itr, school in enumerate(table_as_arrays):

            if school not in skip:
                itr_offset = itr + 3

                # print type(table_as_arrays[itr][0][0])

                ws.cell(row=itr_offset, column=1).value = str(table_as_arrays[itr][0][0])
                ws.cell(row=itr_offset, column=2).value = str(table_as_arrays[itr][1][0])
                ws.cell(row=itr_offset, column=3).value = str(table_as_arrays[itr][2][0])
                ws.cell(row=itr_offset, column=4).value = str(table_as_arrays[itr][3][0])
                ws.cell(row=itr_offset, column=5).value = str(table_as_arrays[itr][4][0])


                done_list.append(school)

                tax_property = models.TaxServicePropertyInformation()
                tax_property.address = str(table_as_arrays[itr][0][0])
                tax_property.description = str(table_as_arrays[itr][1][0])
                tax_property.total_area_m2_unit = str(table_as_arrays[itr][2][0])
                tax_property.price_per_m2_unit = str(table_as_arrays[itr][3][0])
                tax_property.current_rateable_value = str(table_as_arrays[itr][4][0])
                tax_property.billing_authority_code = self.billing_authority_code

                tax_property.billing_authority_link, created = models.BillingAuthority.objects.get_or_create(
                    billing_authority=self.billing_authority_name,
                    billing_authority_code=self.billing_authority_code
                )

                tax_property.building_type, created = models.BuildingType.objects.get_or_create(description=table_as_arrays[itr][1][0])

                tax_property.postcode = tax_property.address.split(',')[-1].strip()

                tax_property.save()

            else:
                print ('Skipped {}'.format(school))

        # print 'done_list', done_list
        return wb.save(filename)

    def clean(self):
        for m in models.TaxServicePropertyInformation.objects.all():
            try:
                # if m.total_area_m2_unit:
                #     m.total_area_m2_unit_num = float(m.total_area_m2_unit.replace(',', ''))
                # if m.current_rateable_value:
                #     m.current_rateable_value_num = float(m.current_rateable_value.replace(',', '').replace('£', ''))

                if m.price_per_m2_unit:
                    m.price_per_m2_unit_num = float(m.price_per_m2_unit.replace(',', '').replace('£', ''))

                m.save()
            except:
                print('Failed:', m.address)

    def find_thing(self, param):
        pubs = models.TaxServicePropertyInformation.objects.filter(building_type__description__contains=param)

        for pub in pubs:
            print(pub)

        print(pubs.count())
        print(pubs.query)

    def record_postcodes(self):
        # postcode_file = 'postcodes.json'

        link_file = '/home/ianh/PycharmProjects/wiserd3/rubbish/beautifulsoup/postcode_lsoa.json'

        with open(link_file, 'r') as link_file:
            links = json.load(link_file)

            print(len(links))
            # print(links[:-10])

            # postcodes = []
            # with open(postcode_file, 'a') as postcode_file:

            errs = 0
            for p in models.TaxServicePropertyInformation.objects.all():

                if not p.lsoa_code:

                    # print(p)
                    # assert isinstance(p, models.TaxServicePropertyInformation)

                    try:
                        link = links[p.postcode]

                        p.lsoa_name = link['lsoa_name']
                        p.lsoa_code = link['lsoa']

                        p.save()
                    except Exception as e:
                        errs += 1
                        # print('err', p.postcode)
                        if len(p.postcode) == 8:
                            pcode_7 = p.postcode.replace(' ', '')

                            try:
                                link = links[pcode_7]

                                p.lsoa_name = link['lsoa_name']
                                p.lsoa_code = link['lsoa']

                                p.save()
                                print('saved', pcode_7)
                            except:
                                print(pcode_7)
            print(errs)
            print(len(links))

                    # postcodes.append(p.postcode)
                # postcode_file.write(json.dumps(postcodes, indent=4))

    def record_geoms(self):
        for prop_info in models.TaxServicePropertyInformation.objects.filter(geocode__isnull=False):
            assert isinstance(prop_info, models.TaxServicePropertyInformation)

            cleaned_postcode = prop_info.postcode
            if len(cleaned_postcode) == 8:
                cleaned_postcode = cleaned_postcode.replace(' ', '')

            postcode_points = dataportal_models.SpatialdataPostCodePoint.objects.filter(postcode=cleaned_postcode)

            if postcode_points.count() > 0:
                prop_info.geom = postcode_points[0].geom
                prop_info.save()

                print('Success {}'.format(cleaned_postcode))

            else:
               print('Failed to find {} ({})'.format(prop_info.postcode, cleaned_postcode))

    def tidy_building_type(self, building_type):
        if building_type:
            building_type = building_type.replace('&', 'and')
            building_type = building_type.replace('and premises', '')
            building_type = building_type.strip()
        return building_type

    def clean_building_types(self, common_name_file='building_fuzz_1.json', unique_name_file='building_fuzz__unique_1.json'):
        from fuzzywuzzy import fuzz

        all_building_types = {}
        seen_building_types = []
        uniques = []

        scrape_building_types = models.BuildingType.objects.all()
        all_count = scrape_building_types.count()

        for building_type_1 in scrape_building_types:
            assert isinstance(building_type_1, models.BuildingType)

            print '{} / {}'.format(building_type_1.id, all_count)
            # We don't want to compare things multiple times
            if building_type_1.id not in seen_building_types:
                building_type_1_text = self.tidy_building_type(
                    building_type_1.description
                )
                building_options = []

                for building_type_2 in scrape_building_types:
                    assert isinstance(building_type_2, models.BuildingType)
                    if building_type_1.id != building_type_2.id:

                        building_type_2_text = self.tidy_building_type(
                            building_type_2.description
                        )
                        score = fuzz.ratio(building_type_1_text, building_type_2_text)

                        if score > 85:
                            seen_building_types.append(building_type_2.id)

                            print(score, building_type_1_text, building_type_2_text)

                            building_options.append(
                                {
                                    'id': building_type_2.id,
                                    'term': building_type_2_text,
                                    'score': score
                                }
                            )
                if len(building_options):
                    all_building_types[building_type_1_text] = building_options
                else:
                    uniques.append({
                        'id': building_type_1.id,
                        'term': building_type_1_text,
                    })

        with open(common_name_file, 'a') as fuzz_file:
            fuzz_file.write(json.dumps(all_building_types, indent=4))

        # Stuff which doesn't match anything
        with open(unique_name_file, 'a') as fuzz_file:
            fuzz_file.write(json.dumps(uniques, indent=4))

    def list_cleaned_building_types(self, common_name_file, unique_name_file, sorted_name_filename):
        name_list = []

        with open(common_name_file, 'r') as fuzz_file:
            fuzz_common_data = json.load(fuzz_file)

            for key in fuzz_common_data:
                name_list.append(key)

        # Stuff which doesn't match anything
        with open(unique_name_file, 'r') as fuzz_unique_file:
            fuzz_unique_data = json.load(fuzz_unique_file)

            for unique_data in fuzz_unique_data:
                name_list.append(unique_data['term'])

        sorted_list = sorted(name_list, key=unicode.lower)

        with open(sorted_name_filename, 'a') as sorted_name_file:
            for n in sorted_list:
                sorted_name_file.write(n + '\n')

    def geocode_address(self, lsoa):
        api_key = 'AIzaSyAxpFEwvBUVjmmDDtAcZYzXci0NyXH7p1Y'

        tspis = models.TaxServicePropertyInformation.objects.filter(
            lsoa_name=lsoa,
            geocode=None,
        )

        tspis_count = tspis.count()
        tspis_itr = 0
        for tspi in tspis:

            tspis_itr += 1
            print '\n{} / {}'.format(tspis_itr, tspis_count)

            assert isinstance(tspi, models.TaxServicePropertyInformation)

            will_continue = False
            if tspi.building_category:
                if tspi.building_category.should_address_point:
                    will_continue = True
                else:
                    will_continue = False
            else:
                will_continue = True

            if will_continue:
                url_address = tspi.address.replace(' ', '+')

                print tspi.address
                # print url_address

                # &bounds=34.172684,-118.604794|34.236144,-118.500938

                url = 'https://maps.googleapis.com/maps/api/geocode/json?address={}&key={}&region=uk'.format(url_address, api_key)

                # print url
                r = requests.get(url)
                print r.status_code
                # print r.text
                json_response = json.loads(r.text)

                if json_response['status'] == 'ZERO_RESULTS' or len(json_response['results']) == 0:
                    print 'error with {}'.format(url_address)
                else:
                    location = json_response['results'][0]['geometry']['location']
                    tspi.geocode = location
                    tspi.save()

    def record_building_categories(self, filter, category, should_address_point):
        tspis = models.TaxServicePropertyInformation.objects.filter(description__istartswith=filter).filter(building_category=None)

        tspis_count = tspis.count()
        tspis_itr = 0
        for tspi in tspis:
            tspis_itr += 1
            print '\n{} / {}'.format(tspis_itr, tspis_count)

            assert isinstance(tspi, models.TaxServicePropertyInformation)
            tspi.building_category, created = models.BuildingCategory.objects.get_or_create(
                description=category,
                should_address_point=should_address_point
            )
            tspi.save()

    def update_geoms(self):
        tspis = models.TaxServicePropertyInformation.objects.filter(geocode__isnull=False)

        tspis_count = tspis.count()
        tspis_itr = 0
        for tspi in tspis:
            assert isinstance(tspi, models.TaxServicePropertyInformation)
            tspis_itr += 1
            print '\n{} / {}'.format(tspis_itr, tspis_count)
            encoded_location = json.loads(tspi.geocode.replace('u\'', '"').replace('\'', '"'))

            print tspi.address
            print encoded_location['lat']
            print encoded_location['lng']

            new_point = Point(encoded_location['lng'], encoded_location['lat'])
            lsoa = dataportal_models.SpatialdataLSOA.objects.get(code=tspi.lsoa_code)

            # We only want points very close to where we expect them to be
            if lsoa.geom.contains(new_point):
                tspi.geom = new_point
                tspi.save()

    def write_xls(self, filename):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'address'
        ws.cell(row=1, column=2).value = 'postcode'
        ws.cell(row=1, column=3).value = 'billing_authority_name'
        ws.cell(row=1, column=4).value = 'description'
        ws.cell(row=1, column=5).value = 'billing_authority_code'
        ws.cell(row=1, column=6).value = 'building_type'
        ws.cell(row=1, column=7).value = 'total_area_m2_unit'
        ws.cell(row=1, column=8).value = 'total_area_m2_unit_num'
        ws.cell(row=1, column=9).value = 'price_per_m2_unit'
        ws.cell(row=1, column=10).value = 'price_per_m2_unit_num'
        ws.cell(row=1, column=11).value = 'current_rateable_value'
        ws.cell(row=1, column=12).value = 'current_rateable_value_num'
        ws.cell(row=1, column=13).value = 'lsoa_name'
        ws.cell(row=1, column=14).value = 'lsoa_code'
        ws.cell(row=1, column=15).value = 'geom_lat'
        ws.cell(row=1, column=16).value = 'geom_lng'

        itr_offset = 3

        tspis = models.TaxServicePropertyInformation.objects.all()[:20]

        for tspi in tspis:
            assert isinstance(tspi, models.TaxServicePropertyInformation)
            # print c_body[key]
            itr_offset += 1

            ws.cell(row=itr_offset, column=1).value = tspi.address
            ws.cell(row=itr_offset, column=2).value = tspi.postcode
            ws.cell(row=itr_offset, column=3).value = tspi.billing_authority_link.billing_authority_name
            ws.cell(row=itr_offset, column=4).value = tspi.description
            ws.cell(row=itr_offset, column=5).value = tspi.billing_authority_code
            ws.cell(row=itr_offset, column=6).value = tspi.building_type.description
            ws.cell(row=itr_offset, column=7).value = tspi.total_area_m2_unit
            ws.cell(row=itr_offset, column=8).value = tspi.total_area_m2_unit_num
            ws.cell(row=itr_offset, column=9).value = tspi.price_per_m2_unit
            ws.cell(row=itr_offset, column=10).value = tspi.price_per_m2_unit_num
            ws.cell(row=itr_offset, column=11).value = tspi.current_rateable_value
            ws.cell(row=itr_offset, column=12).value = tspi.current_rateable_value_num
            ws.cell(row=itr_offset, column=13).value = tspi.lsoa_name
            ws.cell(row=itr_offset, column=14).value = tspi.lsoa_code

            if tspi.geom:
                ws.cell(row=itr_offset, column=15).value = tspi.geom.y
                ws.cell(row=itr_offset, column=16).value = tspi.geom.x

        return wb.save(filename)


if __name__ == "__main__":
    fsm = TaxService()

    billing_authorities = [
        # {
        #     'name': 'Cardiff',
        #     'code': '6815'
        # },

        # {
        #     'code': "6910",
        #     'name': 'Blaenau Gwent'
        # },
        #
        # {
        #     'code': "6915",
        #     'name': 'Bridgend'
        # },
        #
        # {
        #     'code': "6920",pass
        #     'name': 'Caerphilly'
        # },
        #
        # {
        #     'code': "6825",
        #     'name': 'Carmarthenshire(1)'
        # },
        #
        # {
        #     'code': "6828",
        #     'name': 'Carmarthenshire(2)'
        # },
        #
        # {
        #     'code': "6829",
        #     'name': 'Carmarthenshire(3)'
        # },
        #
        # {
        #     'code': "6820",
        #     'name': 'Ceredigion'
        # },
        #
        # {
        #     'code': "6905",
        #     'name': 'Conwy'
        # },

        {
            'code': "6830",
            'name': 'Denbighshire'
        },

        {
            'code': "6835",
            'name': 'Flintshire'
        },

        {
            'code': "6810",
            'name': 'Gwynedd'
        },

        {
            'code': "6805",
            'name': 'Isle of Anglesey'
        },

        {
            'code': "6925",
            'name': 'Merthyr Tydfil'
        },

        {
            'code': "6840",
            'name': 'Monmouthshire'
        },

        {
            'code': "6930",
            'name': 'Neath and Port Talbot'
        },

        {
            'code': "6935",
            'name': 'Newport'
        },

        {
            'code': "6845",
            'name': 'Pembrokeshire'
        },

        {
            'code': "6850",
            'name': 'Powys 1(Montgomeryshire)'
        },

        {
            'code': "6853",
            'name': 'Powys 2(Radnorshire)'
        },

        {
            'code': "6854",
            'name': 'Powys 3(Breconshire)'
        },

        {
            'code': "6940",
            'name': 'Rhondda Cynon Taff'
        },

        {
            'code': "6855",
            'name': 'Swansea'
        },

        {
            'code': "6945",
            'name': 'Torfaen'
        },

        {
            'code': "6950",
            'name': 'Vale of Glamorgan'
        },

        {
            'code': "6955",
            'name': 'Wrexham'
        }
    ]

    # for billing_authority in billing_authorities:
    #     fsm.billing_authority_code = billing_authority['code']
    #     fsm.billing_authority_name = billing_authority['name']
    #     fsm.do_thing()

    # fsm.clean()
    # fsm.find_thing('public')
    # fsm.record_postcodes()
    fsm.record_geoms()
    # fsm.clean_building_types(
    #     common_name_file='building_fuzz_1.json',
    #     unique_name_file='building_fuzz__unique_1.json'
    # )

    # fsm.list_cleaned_building_types(
    #     common_name_file='building_fuzz_1.json',
    #     unique_name_file='building_fuzz__unique_1.json',
    #     sorted_name_filename='sorted_building_types.txt'
    # )

    # fsm.record_building_categories('car park', 'Car Park', False)
    # fsm.record_building_categories('market ', 'Market', False)
    #
    # fsm.geocode_address(lsoa='Cardiff 032F')

    # fsm.update_geoms()
    # fsm.write_xls('dump.xls')
